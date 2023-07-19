from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import openpyxl.styles
from openpyxl.styles import PatternFill, Font
import numpy as np
from tkinter import *
from tkinter import filedialog
from tkinter import ttk
from tkinter import messagebox
import os
import webbrowser
import logging
import shutil


# Configuring Logger

logging.basicConfig(level=logging.DEBUG,
                    filename=r'D:\\log.log',
                    filemode='a',
                    format='%(asctime)s %(name)s %(levelname)s => %(message)s',
                    datefmt="%d-%m-%Y , %H:%M:%S")

#####################  Set Global Variables  ####################

dark_bg = '#0E2954' # Blue
ligth_bg = ''
dark_fg = 'white'
text_fg = "black"
label_font = ("Tahoma", 10)
entry_width = 55

#################################################################
########################  Functionality  ########################
#################################################################



# Fucntion to get the input file path , validate and Show it for the User 
# Enhancement => Try Label Frame ( GUI )

def get_input_file():
    st.set("Status : Listening to Input Path")
    file_input_ent.delete('0',END)
    file_path_i = filedialog.askopenfilename()
    # [01]
    # this if to Check if the User Selected a File or not  
    # if he didn't file_path_i= '' so the insert will through error
    # So this if is to prevent this error from happening 
    if file_path_i != '' :
        file_input_ent.insert(END, file_path_i)
        st.set("Status : Ready")
    else :
        st.set("Status : Input Path is Empty !")
# Fucntion to get the output file path , validate and Show it for the User 

def get_ouput_file():
    # [01]
    st.set("Status : Listening to Output Path")
    file_output_ent.delete('0',END)
    file_path_o = filedialog.asksaveasfilename(defaultextension='.xlsx')
    if file_path_o != '' :
        file_output_ent.insert(END, file_path_o)
        open_file_loc_btn.configure(state=NORMAL)
        st.set("Status : Ready")
    else :
        st.set("Status : Output Path is Empty !")




# Validate All Entieries From The User
def check_user_input() :
    global rl ; global input_file ; global output_file
    try :
        input_file = file_input_ent.get()
        output_file = file_output_ent.get()
        rl = bm_ent.get().strip()
        st.set("Status : Checking User Input")
        if input_file == '':
            st.set("Status : Warning => Input File Path is Empty !")
            messagebox.showwarning("Error", "Specify Input File Path First !")
            raise ValueError("User Didn't Choose Input Path")
        elif not os.path.exists(input_file):
            st.set("Status : Warning => Input Path is Not Valid")
            messagebox.showwarning("Error", "Input Path Doesn't Exist")
            file_input_ent.delete('0', END)
            raise ValueError("Input Path is Not Valid")
        elif output_file == '':
            st.set("Status : Warning => Output File Path is Empty !")
            messagebox.showwarning("Error", "Specify Output File Path First !")
            raise ValueError("User Didn't Choose Output Path")
        elif not os.path.exists(os.path.dirname(output_file)):
            st.set("Status : Warning => Output Path is Not Valid")
            messagebox.showwarning("Error", "Output Path Doesn't Exist")
            file_output_ent.delete('0', END)
            raise ValueError("Output Path is Not Valid")
        elif rl == '':
            st.set("Status : Warning => B.M Level is Empty")
            messagebox.showwarning("Error", "Input B.M Level First !")
            raise ValueError("B.M Field Is Empty")
        elif selected_value.get() == '' :
            st.set("Status : Warning => Select Number of Reading First !")
            messagebox.showwarning("Error","Please Select Number of Readings")
            raise ValueError("User Didn't Choose (1 or 3) Stadia Readings")
        elif not rl == '' :
            try :
                float(rl)
                logging.info("Try => True")
            except Exception :
                messagebox.showerror(f"Error With '{rl}'","R.L Must be Digit Not Charachter")
                raise ValueError("R.L Must be int() or Float() not Charachter")
        st.set("Status : User Input is Valid")
        logging.debug("Validation of User Input Were Done Succesfully")
    except Exception as e :
        logging.error(f"Error With 'check_user_input()' => {e}")
        # To Prevent the Next Function from Run in Try Except Block of Calculate Function
        raise ValueError("Execution of the Code Stopped => Loop") 



def handle_excel_file() :
    try :
        st.set('Status : Reading Excell File .... ')
        global wb ; global ws ; global rows_num ; global cols_num ; global max_row
        wb = load_workbook(input_file)
        # Selcet Work Sheet
        ws = wb.active
        rows_num = ws.max_row
        cols_num = ws.max_column
        max_row = ws.max_row
        logging.info("Excel File Accessed Successfully")
        st.set('Status : Excell File Read Successfully ')
    except Exception as e :
        name1 = os.path.basename(input_file)
        st.set(f"Status : Error => Can't read {name1} , May Be it's not Excell Sheet")
        logging.error(f"Error With 'handle_excel_file()' => {e}")
        raise ValueError("Execution of the Code Stopped => Loop")

def check_excell_file() :
    try :
        st.set('Status : Checking Excell File .... ')
        if type(ws['A1'].value) == str or bool(ws['A1'].value) == False :
                messagebox.showerror("Error",'Unsupported File Format !')
                st.set('Status : Error => The File May Contain a Header or the Row Number 1 is Empty ! ')
                mm = 'Unsupported File Format , The File May Contain a Header or the Row Number 1 is Empty !'
                raise ValueError(mm)
        else:
                logging.info("check_excell_file => True")
                st.set('Status : Excell File Checked Succesfully')
    except Exception as e :
            logging.error(f"Error With 'handle_excel_file()' => {e}")
            raise ValueError("Execution of the Code Stopped => Loop")

def check_excell_file3() :
    try :
        if type(ws['A1'].value) == str or type(ws['A2'].value) == str or type(ws['A3'].value) == str or bool(ws['A1'].value) == False or bool(ws['A2'].value) == False or bool(ws['A3'].value) == False  :
            st.set('Status : Error => The File May Contain a Header or the Row Number 1 is Empty ! ')
            messagebox.showerror("Error",'Unsupported File Format !')
            mm = 'Unsupported File Format , The File May Contain a Header or the Row Number 1 is Empty !'
            raise ValueError(mm)
        else :
            logging.info("check_excell_file3 => True")
    except Exception as e:
            logging.error(f"Error With 'handle_excel_file3()' => {e}")
            raise ValueError("Execution of the Code Stopped => Loop")


def check_none_values() :
    st.set('Status : Checking None Values ..... ')
    try :
        # Check None Values
        for r in range(1, rows_num + 1):
            for c in range(1, cols_num + 1):
                if ws[f"{get_column_letter(c)}{r}"].value == None:
                    ws[f"{get_column_letter(c)}{r}"] = 0
        logging.info("None Value Checked Succefully")
    except Exception as e :
        st.set('Status : Error => None Value Checking Failed')
        logging.error(f"Error With 'handle_excel_file3()' => {e}")

# chack readings 
def check_readings() :
    st.set('Status : Start Check Readings .....')
    global BS ; global IS ; global FS
    try :
        # Define BS | IS | FS Objects
        bsr = ws['A1':f'A{rows_num}']
        isr = ws['B1':f'B{rows_num}']
        fsr = ws['C1':f'C{rows_num}']

        BS = []
        IS = []
        FS = []

        # Get Values From Objects and Append it to the list
        for i in range(0,rows_num):
            BS.append(bsr[i][0].value)
            IS.append(isr[i][0].value)
            FS.append(fsr[i][0].value)
        
        logging.info("BS | IS | FS Read Successfully")
        r = 1
        if BS[0] !=0 and IS[0] ==0 and FS[0] == 0 :
            logging.info(f"Checking Row Number {r} => OK")
        else :
            logging.error(f"Checking Row Number {r} => Error ")
            m = f"Check Row Number : {r}"
            messagebox.showerror('Error',f"Error in Readings , {m} ")
            raise ValueError(f'Error in Row Number {r}')
        
        # Check in between Rowss
        for i in range(2,rows_num) :
            r+=1
            if BS[i-1] !=0 and IS[i-1] ==0 and FS[i-1] != 0 :
                    logging.info(f"Checking Row Number {r} => OK")
            elif BS[i-1] == 0 and IS[i-1] !=0 and FS[i-1] == 0 :
                logging.info(f"Checking Row Number {r} => OK")
            else :
                logging.error(f"Checking Row Number {r} => Error")
                m = f"Check Row Number : {r}"
                messagebox.showerror('Error',f"Error in Readings , {m} ")
                raise ValueError(f'Error in Row Number {r}')

        # Check Last Row 
        if BS[rows_num-1] ==0 and IS[rows_num-1] ==0 and FS[rows_num-1] != 0 :
            logging.info(f"Checking Row Number {rows_num+1} => OK")
        else :
            logging.error(f"Checking Row Number {rows_num+1} => Error")
            m = f"Check Row Number : {rows_num+1}"
            messagebox.showerror('Error',f"Error in Readings , {m} ")
            raise ValueError(f'Error in Row Number {rows_num+1}')
        logging.info("All Readings Checked Successfully")
    except Exception as e  :
        st.set('Status : Error =>  Unsupported File Format ,"check_readings()"')
        messagebox.showerror("Error","Unsupported File Format ")
        logging.error(f"Error Occured With 'check_readings()' => {e} ")
        raise ValueError("Execution of the Code Stopped => Loop")

def rl_calculations():
    st.set('Status : Start R.L Calculations ....')
    global rounded_reduced_level2
    rl = float(bm_ent.get())
    try : 
        reduced_level = [rl]
        hi = BS[0] + rl
        for i in range(1, rows_num):
            if IS[i] != 0:
                val = hi - IS[i]
                reduced_level.append(val)
                hi = hi
            elif FS != 0:
                val = hi - FS[i]
                reduced_level.append(val)
                hi = val + BS[i]
            rounded_reduced_level2 = [round(num, 3) for num in reduced_level]
        logging.info("R.L Calculation Done Succefully")
    except Exception as e :
            st.set("Status : Error => Can't Calculate R.L")
            logging.error(f"Error Occured With 'rl_calculations()' => {e} ")
            raise ValueError("Execution of the Code Stopped => Loop")


def insert_data() :
    st.set('Status : Inserting Data To Execll Sheet .....')
    try :
        # Insert Header
        head = ["NO.", "BS", "IS", "FS", "RL"]
        for n in range(1, 6):
            ws.cell(row=1, column=n, value=head[n - 1])
        logging.info("Header Created Succefully")

        # Insert Point Number | BS | IS | FS | RL in the Excell sheet
        array1 = np.array([range(1, rows_num + 1), BS, IS, FS, rounded_reduced_level2]).transpose()
        for i in range(2, rows_num + 2):
            for j in range(1, cols_num + 3):
                ws.cell(row=i, column=j, value=array1[i - 2, j - 1])
        logging.info("Data Inserted into Exell Sheet Succefully")
    except Exception as e :
        st.set("Status : Error => Can't Insert Data")
        logging.error(f"Error Occured With 'insert_data()' => {e} ")
        raise ValueError("Execution of the Code Stopped => Loop")


def foramtting() :
    st.set('Status : Start Formatting .....')
    try :
        # Set the alignment for all cells to center
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        logging.info("Alignment => Center done Succefully")

        # Define the font and fill for the header style
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

        # Set the header style for the first row
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
        logging.info("File Formatted Succesfully")
        # Save Changes to Specific Path
        wb.save(output_file)
        logging.info(f"File Was Saved to {output_file} Succesfully")

        # Declare Complete For User 
        st.set('Status : Formating Success')
        messagebox.showinfo("Complete", "Calculation Completed and File Was Saved Successfully".title())

        # Open File Button
        open_file_btn.configure(state=NORMAL)
        
        # Shutdown logger
        try :
            src_path = r'D:\\log.log'
            dist_path = os.path.dirname(file_output_ent.get())
            logging.info("Logging File Saved Succefully")
            logging.info("Logger Shutdown")
            shutil.move(src_path,dist_path)
            logging.shutdown()
        except Exception as e :
            logging.warning("File Already Exists {log}, Saved to Existing File")
            st.set('Status :  Completed')
    except Exception as e :
        st.set('Status : Formating Failed')
        logging.error(f"Error Occured With 'formatting()' => {e} ")
        raise ValueError("Execution of the Code Stopped => Loop")


#################################################################
#########################    3 Reading  #########################
#################################################################

def read_BIF3():
    try :
        global u_bs , u_is , u_fs , m_bs , m_is , m_fs , l_bs ,l_is ,l_fs , bs_val ,is_val,fs_val
        
        # Container of BS readings
        u_bs = []
        m_bs = []
        l_bs = []

        # Access BS column {objects}
        bs_col = ws["A1":f"A{max_row}"]

        # Convert BS objects to Values
        bs_val = []
        for cell in range(0,max_row) :
            bs_val.append(bs_col[cell][0].value)

        # Get Values of Upper , Middle , Lower BS
        for cell in range(0,max_row,3) :
            u_bs.append(bs_col[cell][0].value)
            m_bs.append(bs_col[cell+1][0].value)
            l_bs.append(bs_col[cell+2][0].value)
        
        logging.info("U.M.L Readings B.S Read Successfully")

        # Container of IS readings
        u_is = []
        m_is = []
        l_is = []

        # Access IS column {objects}
        is_col = ws["B1":f"B{max_row}"]

        # Convert IS objects to Values
        is_val = []
        for cell in range(0,max_row) :
            is_val.append(is_col[cell][0].value)


        # Get Values of Upper , Middle , Lower IS
        for cell in range(0,max_row,3) :
            u_is.append(is_col[cell][0].value)
            m_is.append(is_col[cell+1][0].value)
            l_is.append(is_col[cell+2][0].value)
        
        logging.info("U.M.L Readings I.S Read Successfully")
        # Container of FS readings
        u_fs = []
        m_fs = []
        l_fs = []

        # Access FS column {objects}
        fs_col = ws["C1":f"C{max_row}"]

        # Convert FS objects to Values
        fs_val = []
        for cell in range(0,max_row) :
            fs_val.append(fs_col[cell][0].value)
        
        # Get Values of Upper , Middle , Lower FS
        for cell in range(0,max_row,3) :
            u_fs.append(fs_col[cell][0].value)
            m_fs.append(fs_col[cell+1][0].value)
            l_fs.append(fs_col[cell+2][0].value)
        
        logging.info("U.M.L Readings F.S Read Successfully")
    except Exception as e :
        st.set("Status : Error3 => Can't Read Data 'read_BIF3()'")
        messagebox.showerror("Error with 'read_BIF3()'",f"{e}")
        logging.error(f"Error Occured With 'read_BIF3()' => {e} ")
        raise ValueError("Execution of the Code Stopped => Loop")

def check_readings3() :

    try :
        # Check Row 1 
        r = 1

        if u_bs[0] !=0 and u_is[0] == 0 and u_fs[0] == 0 and m_bs[0] !=0 and m_is[0] == 0 and m_fs[0] == 0 and l_bs[0] !=0 and l_is[0] == 0 and l_fs[0] == 0  :
            logging.info(f"Checking Row Number {r} => OK")
        else :
            logging.error(f"Checking Row Number {r}  => Error ")
            m = f"Check Row Number : {r}"
            logging.error('Error',f"Error in Readings , {m} ")

        # Check in between rows
        for n in range(1,len(u_bs)-1):
            r+=1
            if  u_bs[n] == 0 and u_is[n] != 0 and u_fs[n] == 0 and m_bs[n] ==0 and m_is[n] != 0 and m_fs[n] == 0 and l_bs[n] ==0 and l_is[n] != 0 and l_fs[n] == 0  :
                logging.info(f"Checking Row Number {r} => OK")
            elif u_bs[n] !=0 and u_is[n] == 0 and u_fs[n] != 0 and m_bs[n] !=0 and m_is[n] == 0 and m_fs[n] != 0 and l_bs[n] !=0 and l_is[n] == 0 and l_fs[n] != 0  : 
                logging.info(f"Checking Row Number {r} => OK")
            else :
                logging.error(f"Checking Row Number {r}  => Error ")
                m = f"Check Row Number : {r}"
                logging.error('Error',f"Error in Readings , {m} ")
        
        # Check Last Row 

        z = len(u_bs)-1

        if u_bs[z] ==0 and u_is[z] == 0 and u_fs[z] != 0 and m_bs[z] ==0 and m_is[z] == 0 and m_fs[z] != 0 and l_bs[z] ==0 and l_is[z] == 0 and l_fs[z] != 0  :
            logging.info(f"Checking Row Number {z+1} => OK")
        else :
            logging.error(f"Checking Row Number {z+1}  => Error ")
            m = f"Check Row Number : {z+1}"
            logging.error('Error',f"Error in Readings , {m} ")
        logging.info("All Readings Checked Succefully")

    except Exception as e :
        st.set('Status : Error3 => Readings Check Failed !')
        messagebox.showerror("Error with 'read_BIF3()'",f"{e}")
        logging.error(f"Error Occured With 'read_BIF3()' => {e} ")
        raise ValueError("Execution of the Code Stopped => Loop")


# RL Calculations 
def rl_calculations3():
    try:
        global r_l , rounded_reduced_level
        rl = float(bm_ent.get())
        reduced_level = [rl]
        hi = m_bs[0] + rl
        for i in range(1, len(u_bs)):
            if m_is[i] != 0:
                val = hi - m_is[i]
                reduced_level.append(val)
                hi = hi
            elif m_fs != 0:
                val = hi - m_fs[i]
                reduced_level.append(val)
                hi = val + m_bs[i]
            rounded_reduced_level = [round(num, 3) for num in reduced_level]
        logging.info("Calculation of R.L Done Succefully")
        # for loop to set the calculated r_l in their correct position
        r_l = [0]*max_row
        x = 0 
        for i in range(1,max_row,3):
            if x ==len(u_bs) :
                break
            else :
                r_l[i] = rounded_reduced_level[x]
                x += 1
        logging.info("R.L Indexing Success")
    except Exception as e:
        st.set("Status : Error3 => Can't Calculate R.L ")
        messagebox.showerror("Error with 'rl_calculations3()'",f"{e}")
        logging.error(f"Error Occured With 'rl_calculations3()' => {e} ")
        raise ValueError("Execution of the Code Stopped => Loop")


def calculate_distance() :
    try :
        global dist
        distance = []

        for i in range(0, len(u_bs)):
            if u_bs[i] !=0 and u_fs[i] != 0 :
                # we will calculated two distances 
                d1 = str(round(((u_bs[i] - l_bs[i]) * 100),3))
                d2 = str(round(((u_fs[i] - l_fs[i]) * 100),3))
                distance.append(f"({d2},{d1})")
            elif u_bs[i] != 0 and u_is[i] == 0 and u_fs[i] ==0 :
                distance.append(round(( u_bs[i] - l_bs[i] ) *100,3))
            elif u_bs[i] == 0 and u_is[i] != 0 and u_fs[i] ==0 :
                distance.append(round(( u_is[i] - l_is[i] ) *100,3))
            elif u_bs[i] == 0 and u_is[i] == 0 and u_fs[i] !=0 :
                distance.append(round(( u_fs[i] - l_fs[i] ) *100,3))

        logging.info("Calculation of Distances Done Succefully")
        # for loop to set the calculated r_l in their correct position
        dist = [0]*max_row
        x = 0 
        for i in range(1,max_row,3):
            if x == 6 :
                break
            else :
                dist[i] = distance[x]
                x += 1
        logging.info("Distances Indexing Success")
    except Exception as e :
        st.set("Status : Error3 => Can't Calculated Distances ")
        messagebox.showerror("Error with 'calculate_distance()'",f"{e}")
        logging.error(f"Error Occured With 'calculate_distance()' => {e} ")
        raise ValueError("Execution of the Code Stopped => Loop")




# Insert Data 

def insert_data3() :
    try :
        No = range(1, max_row + 1)
        head = ["NO.", "BS", "IS", "FS", "RL","DIST"]

        # Concatnation
        for n in range(1, 7):
            ws.cell(row=1, column=n, value=head[n - 1])
        
        logging.info("Header Created Successfully")
        array1 = np.array([No,bs_val,is_val,fs_val,r_l]).transpose()
        for i in range(2, max_row + 2):
            for j in range(1, 6):
                ws.cell(row=i, column=j, value=array1[i - 2, j - 1])
        logging.info("Data Inserted into Exell Sheet Succefully")

        for r in range(2,max_row+2) :
            ws[f'F{r}'] = dist[r-2]
        logging.info("Distances Insertion Success")
    except Exception as e :
        st.set('Status : Error3 => Data Insertion Failed')
        messagebox.showerror("Error with 'insert_data3()'",f"{e}")
        logging.error(f"Error Occured With 'insert_data3()' => {e} ")
        raise ValueError("Execution of the Code Stopped => Loop")




def get_report_data3():
    global data3
    try :
        data3 = []
        total_bs = sum(m_bs)
        total_fs = sum(m_fs)
        a = "=========================================================================\n"
        b = f"Applying The Condition Number of BS = Number of FS\n"
        q = rounded_reduced_level
        if len(m_bs) == len(m_fs) :
            w = f"(No. BS = {len(m_bs)}, No. FS = {len(m_fs)}) => True\n"
        else :
            w = f"(No. BS = {len(m_bs)}, No. FS = {len(m_fs)}) => False\n"

        # rounded last - first RL
        r_l_f_rl = round(q[-1]-q[0],3)

        # rounded sum BS - sum FS
        ff = round((total_bs)-(total_fs),3)

        # rounded total bs

        s1 = round(total_bs,3)

        # rounded total fs
        s2 = round(total_fs,3)

        c = "Applying The Arithmetic Check => sum(BS) - sum(FS) = Last RL - First RL\n"
        if round((total_bs-total_fs),3) == round(q[-1]-q[0],3) :
            d = f"sum(BS) - sum(FS)\n{s1} - {s2} = {ff}\n" 
            e = f"Last RL - First RL\n{q[-1]} - {q[0]} = {r_l_f_rl}\n"
            x = f"{ff} = {r_l_f_rl}  => True\n"
        else:
            d = f"sum(BS) - sum(FS)\n{s1} - {s2} = {ff}\n" 
            e = f"Last RL - First RL\n{q[-1]} - {q[0]} = {r_l_f_rl}\n"
            x = f"{ff} = {r_l_f_rl}  => False\n"
        f = "=========================================================================\n"
        data3.append(a)
        data3.append(b)
        data3.append(w)
        data3.append(c)
        data3.append(d)
        data3.append(e)
        data3.append(x)
        data3.append(f)
        
    except Exception as e :
        st.set('Status : Error3 => Cannot Generate Report Data')
        messagebox.showerror("Error","Error With 'get_report_data3()'")
        logging.error(f"Error With 'get_report_data3() => {e}")
        raise ValueError("Execution of the Code Stopped => Loop")

def get_report_data():
    global data3
    try :
        data3 = []
        total_bs = sum(BS)
        total_fs = sum(FS)
        a = "=========================================================================\n"
        b = f"Applying The Condition Number of BS = Number of FS\n"
        q = rounded_reduced_level2
        if len(BS) == len(FS) :
            w = f"(No. BS = {len(BS)}, No. FS = {len(FS)}) => True\n"
        else :
            w = f"(No. BS = {len(BS)}, No. FS = {len(FS)}) => False\n"

        # rounded last - first RL
        r_l_f_rl = round(q[-1]-q[0],3)

        # rounded sum BS - sum FS
        ff = round((total_bs)-(total_fs),3)

        # rounded total bs

        s1 = round(total_bs,3)

        # rounded total fs
        s2 = round(total_fs,3)

        c = "Applying The Arithmetic Check => sum(BS) - sum(FS) = Last RL - First RL\n"
        if round((total_bs-total_fs),3) == round(q[-1]-q[0],3) :
            d = f"sum(BS) - sum(FS)\n{s1} - {s2} = {ff}\n" 
            e = f"Last RL - First RL\n{q[-1]} - {q[0]} = {r_l_f_rl}\n"
            x = f"{ff} = {r_l_f_rl}  => True\n"
        else:
            d = f"sum(BS) - sum(FS)\n{s1} - {s2} = {ff}\n" 
            e = f"Last RL - First RL\n{q[-1]} - {q[0]} = {r_l_f_rl}\n"
            x = f"{ff} = {r_l_f_rl}  => False\n"
        f = "=========================================================================\n"
        data3.append(a)
        data3.append(b)
        data3.append(w)
        data3.append(c)
        data3.append(d)
        data3.append(e)
        data3.append(x)
        data3.append(f)
    except Exception as e :
        st.set('Status : Error => Cannot Generate Report Data')
        messagebox.showerror("Error","Error With 'get_report_data()'")
        logging.error(f"Error With 'get_report_data3() => {e}")
        raise ValueError("Execution of the Code Stopped => Loop")


def generate_report() :
    try :
        # Ask User For the path name
        path = filedialog.asksaveasfilename(defaultextension='txt')
        name = os.path.basename(path)
        # Get Path Without File Name
        dir = os.path.dirname(path)
        report_file = open(file=path,mode='a')
        report_file.writelines(data3)
        messagebox.showinfo("Complete",f"Report Genetrated Succesfully to => {path}")
    except Exception as e :
        st.set('Status : Error => Occured While Writing Report ...')
        messagebox.showerror("Error","Error With 'generate_report()'")
        logging.error(f"Error With 'generate_report()' => {e}")
        raise ValueError("Execution of the Code Stopped => Loop")

#################################################################
#########################    End  ###############################
#################################################################


def open_folder():
    global dir_path
    # Get the directory path
    dir_path = os.path.dirname(file_output_ent.get())
    # Open The folder
    os.startfile(dir_path)


def open_file():
    os.startfile(file_output_ent.get())

def clear():
    file_input_ent.delete('0', END)
    file_output_ent.delete('0', END)
    bm_ent.delete('0', END)
    open_file_btn.configure(state=DISABLED)
    open_file_loc_btn.configure(state=DISABLED)
    st.set('Status : Ready')
    selected_value.set('')
    report_btn.configure(state=DISABLED)

# Dark Mode
def dark():
    root.configure(bg=dark_bg)
    dark_m_btn.configure(command=light)
    dark_m_btn['text'] = "Light Mode"
    contact.configure(bg=dark_bg, fg=dark_fg)
    get_data.configure(bg=dark_bg,fg=dark_fg)
    file_input_label.configure(bg=dark_bg, fg=dark_fg)
    file_output_label.configure(bg=dark_bg, fg=dark_fg)
    bm_label.configure(bg=dark_bg, fg=dark_fg)
    fb_btn.configure(bg=dark_bg, activebackground=dark_bg)
    te_btn.configure(bg=dark_bg, activebackground=dark_bg)
    wt_btn.configure(bg=dark_bg, activebackground=dark_bg)
    ge_btn.configure(bg=dark_bg, activebackground=dark_bg)
    footer.configure(bg='#84A7A1', fg='white', font=('Trebuchet MS', 10))
    statue.configure(bg='white')
    frame.configure(bg=dark_bg)
    statue.configure(bg='white')
    style.configure("TRadiobutton", background=dark_bg, foreground = dark_fg)

# Light Mode
def light():
    root.configure(bg=root_bg)
    dark_m_btn.configure(command=dark)
    dark_m_btn['text'] = "Dark Mode"
    contact.configure(bg=root_bg, fg=text_fg)
    get_data.configure(bg=root_bg,fg=text_fg)
    file_input_label.configure(bg=root_bg, fg=text_fg)
    file_output_label.configure(bg=root_bg, fg=text_fg)
    bm_label.configure(bg=root_bg, fg=text_fg)
    fb_btn.configure(bg=root_bg, activebackground=root_bg)
    te_btn.configure(bg=root_bg, activebackground=root_bg)
    wt_btn.configure(bg=root_bg, activebackground=root_bg)
    ge_btn.configure(bg=root_bg, activebackground=root_bg)
    footer.configure(bg='white', fg='black', font=('Trebuchet MS', 10))
    statue.configure(bg='white')
    frame.configure(bg=root_bg)
    statue.configure(bg='#b8c3d4')
    style.configure("TRadiobutton", background=root_bg ,  foreground=text_fg)

def calculate() :
    try :
        check_user_input()
        handle_excel_file()
        if selected_value.get() ==  "1":
            check_excell_file()
        elif selected_value.get() == "2" :
            check_excell_file3()
        check_none_values()
        if selected_value.get() ==  "1":
            check_readings()
            rl_calculations()
            insert_data()
            get_report_data()
            foramtting()
            report_btn.configure(state=NORMAL)
        elif selected_value.get() == "2" :
            read_BIF3()
            check_readings3()
            rl_calculations3()
            calculate_distance()
            insert_data3()
            get_report_data3()
            foramtting()
            report_btn.configure(state=NORMAL)
    except Exception as e :
        logging.error(f"Error With 'calculate()' => {e}")

# Contact Me Buttons

f = "https://www.facebook.com/es.sa.18659/"
t = "https://t.me/ahmedayman105"
wa = 'https://wa.me/+201099605975'
g = 'mailto:ahmedaymanissa98@gmail.com'

def open_f():
    webbrowser.open_new_tab(f)

def open_t():
    webbrowser.open_new_tab(t)

def open_w():
    webbrowser.open_new_tab(wa)

def open_g():
    webbrowser.open_new_tab(g)

#################################################################
#########################     GUI      ##########################
#################################################################

#############################  Root  ############################

root = Tk()
root.title("Survey App")
root.iconbitmap(r"images\logo.ico")
root_bg = "#F5F5F5"
root.configure(bg=root_bg)
root.resizable(False,False)


########################  Padding Frame  ########################

frame = Frame(root,bg=root_bg)
frame.pack()

#########################  Label Frame  #########################

get_data = LabelFrame(frame,text=' Leveling Calculator ',bg=root_bg,font=('Tahoma',8,'bold'))
get_data.grid(row=0,column=0,padx=10,pady=10,ipadx=5,ipady=5)

############################  Row 1  #############################

file_input_label = Label(get_data, text="Choose File", bg=root_bg, fg=text_fg, font=label_font)
file_input_label.grid(row=0,column=0)

file_input_ent = ttk.Entry(get_data, width=entry_width)
file_input_ent.grid(row=0,column=1)

file_input_btn = ttk.Button(get_data, text="Browse",command=get_input_file)
file_input_btn.grid(row=0,column=2)

############################  Row 2  #############################

file_output_label = Label(get_data, text="Output Path", bg=root_bg, fg=text_fg, font=label_font)
file_output_label.grid(row=1,column=0)

file_output_ent = ttk.Entry(get_data, width=entry_width)
file_output_ent.grid(row=1,column=1)

file_output_btn = ttk.Button(get_data, text="Browse",command=get_ouput_file)
file_output_btn.grid(row=1,column=2)

############################  Row 3  #############################

bm_label = Label(get_data, text="B.M Level", bg=root_bg, fg=text_fg, font=label_font)
bm_label.grid(row=2,column=0,sticky=W)

bm_ent = ttk.Entry(get_data, width=entry_width)
bm_ent.grid(row=2,column=1)

report_btn = ttk.Button(get_data,text='Report',state=DISABLED,command=generate_report)
report_btn.grid(row=2,column=2)

bm_btn = ttk.Button(get_data, text="Calculate",command=calculate)
bm_btn.grid(row=3,column=2)

##############  Set Padding For All Childs   ######################

for w in get_data.winfo_children():
    w.grid_configure(padx=15,pady=10)

############################  Row 4  #############################


style = ttk.Style()
style.configure("TRadiobutton", background=root_bg)

selected_value = StringVar(value='')

one_reading = ttk.Radiobutton(get_data,text='1 Reading',value=1,style="TRadiobutton",variable=selected_value)
one_reading.grid(row=3,column=1,padx=50,pady=10,sticky=W)

three_reading = ttk.Radiobutton(get_data,text='3 Readings',value=2,style="TRadiobutton",variable=selected_value)


three_reading.grid(row=3,column=1,padx=50,pady=10,sticky=E)


############################  Row 5  #############################


open_file_loc_btn = ttk.Button(get_data, text="Open File Location", state=DISABLED,command=open_folder)
open_file_loc_btn.grid(row=4,column=1,padx=15,sticky=W)

open_file_btn = ttk.Button(get_data, text="Open File", state=DISABLED , command=open_file)
open_file_btn.grid(row=4,column=1)

dark_m_btn = ttk.Button(get_data, text="Dark Mode",command=dark)
dark_m_btn.grid(row=4,column=1,sticky=E,padx=15)

clear_btn = ttk.Button(get_data, text="Clear",command=clear)
clear_btn.grid(row=4,column=2,pady=10)



#######################  Contact Me Frame  #######################

contact = LabelFrame(frame,text=' Contact Me ',bg=root_bg,font=('Tahoma',8,'bold'))
contact.grid(row=1,column=0,padx=10,pady=(10,100),ipadx=5,ipady=5,sticky=EW)


c = 'hand2'

# Facebook Button

# Icon
facebook = PhotoImage(file=r"images\facebook2424.png")

# Button
fb_btn = Button(contact, image=facebook, borderwidth=0, bg=root_bg, activebackground=root_bg, cursor=c,command=open_f)
fb_btn.grid(row=0,column=0)


# Telegram Button

# Icon
telegram = PhotoImage(file=r"images\telegram2424.png")

# Button

te_btn = Button(contact, image=telegram, borderwidth=0, bg=root_bg, activebackground=root_bg, cursor=c,command=open_t)
te_btn.grid(row=0,column=1)
# te_btn.place(x=x+50, y=y, anchor=anchor_w)


# Whatsapp Button

# Icon
whatsapp = PhotoImage(file=r"images\whatsapp2424.png")

# Button
wt_btn = Button(contact, image=whatsapp, borderwidth=0, bg=root_bg, activebackground=root_bg, cursor=c,command=open_w)
wt_btn.grid(row=0,column=2)


# Gmail Button

# Icon
gmail = PhotoImage(file=r"images\gmail2424.png")

# Button
ge_btn = Button(contact, image=gmail, borderwidth=0, bg=root_bg, activebackground=root_bg, cursor=c,command=open_g)
ge_btn.grid(row=0,column=3)

for w in contact.winfo_children() :
    w.grid_configure(padx=20,pady=10)

contact.grid_columnconfigure(0, weight=1)
contact.grid_columnconfigure(1, weight=1)
contact.grid_columnconfigure(2, weight=1)
contact.grid_columnconfigure(3, weight=1)

#########################  Status Bar   ########################

st = StringVar(value='Status : Ready')
statue = Label(root,bg="#b8c3d4",font=('Tahoma',10),anchor='w',textvariable=st)
statue.place(rely=1,y=-40,anchor='sw',relx=0,relwidth=1)


############################  Footer  ###########################

footer = Label(root, text="Powered By Ahmed Ayman © 2023", height=2, bg="white", font=('Tahoma', 10))
footer.place(rely=1, relwidth=1, anchor='sw')


############################  Loop  #############################

root.mainloop()